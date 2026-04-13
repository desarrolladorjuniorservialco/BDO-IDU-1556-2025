"""
sync_contrato.py — Seguimiento contractual · BDO IDU-1556-2025

Lee Contrato_IDU_1556_2025.xlsx una sola vez y sincroniza:
  · contratos          ← hoja BD_CTO_INI
  · contratos_prorrogas ← hoja BD_CTO_PRO
  · contratos_adiciones ← hoja BD_CTO_ADI

Los contadores prorrogas/adiciones y los valores plazo_actual/valor_actual
de la tabla contratos son mantenidos automáticamente por triggers de BD;
esta función sólo upserta los datos base y el detalle de cada hoja.
"""
import openpyxl
from datetime import datetime, date

from .config import CONTRATO_ID
from .gpkg import download_file
from .utils import safe

EXCEL_CONTRATO = 'Contrato_IDU_1556_2025.xlsx'
TMP_PATH       = '/tmp/contrato.xlsx'


# ── helpers ────────────────────────────────────────────────────────────────

def _to_date(val):
    """Convierte datetime/date de openpyxl a 'YYYY-MM-DD', o devuelve None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    return safe(val)


def _to_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _sheet_rows(ws):
    """
    Devuelve lista de dicts {encabezado: valor} ignorando filas vacías.
    La primera fila no vacía es el encabezado.
    """
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    if len(rows) < 2:
        return []
    headers = [
        str(h).strip().lower() if h is not None else f'_col{i}'
        for i, h in enumerate(rows[0])
    ]
    return [dict(zip(headers, row)) for row in rows[1:]]


def _load_wb(token, project_id):
    """Descarga el Excel y retorna el workbook, o None si falla."""
    if not download_file(token, project_id, EXCEL_CONTRATO, TMP_PATH):
        return None
    try:
        return openpyxl.load_workbook(TMP_PATH, data_only=True)
    except Exception as e:
        print(f"  ✗ Error leyendo Excel: {e}")
        return None


# ── funciones públicas ──────────────────────────────────────────────────────

def sync_contrato_excel(supabase, token, project_id):
    """
    Punto de entrada único: descarga el Excel una vez y sincroniza
    las tres hojas contractuales en orden (INI → PRO → ADI).
    """
    print("\n── Contrato Excel ──")
    wb = _load_wb(token, project_id)
    if wb is None:
        return

    _sync_ini(supabase, wb)
    _sync_pro(supabase, wb)
    _sync_adi(supabase, wb)


# ── internos ────────────────────────────────────────────────────────────────

def _sync_ini(supabase, wb):
    """Sincroniza contratos desde hoja BD_CTO_INI."""
    print("  · BD_CTO_INI → contratos")
    if 'BD_CTO_INI' not in wb.sheetnames:
        print("  ⚠ Hoja BD_CTO_INI no encontrada")
        return

    count = errores = 0
    for row in _sheet_rows(wb['BD_CTO_INI']):
        cid = safe(row.get('id'))
        if not cid:
            continue
        data = {
            'id':            cid,
            'nombre':        safe(row.get('nombre')),
            'contratista':   safe(row.get('contratista')),
            # El Excel usa 'intrventoria' (typo deliberado, coincide con BD)
            'intrventoria':  safe(row.get('intrventoria') or row.get('interventoria')),
            'supervisor_idu': safe(row.get('supervisor_idu')),
            'fecha_inicio':  _to_date(row.get('fecha_inicio')),
            'fecha_fin':     _to_date(row.get('fecha_fin')),
            'valor_contrato': _to_int(row.get('valor_contrato')),
            # prorrogas/adiciones/plazo_actual/valor_actual los mantiene el trigger
        }
        data = {k: v for k, v in data.items() if v is not None}
        try:
            supabase.table('contratos').upsert(data, on_conflict='id').execute()
            count += 1
            print(f"    ✓ {cid}")
        except Exception as e:
            errores += 1
            print(f"    ✗ {cid}: {e}")

    print(f"    → {count} upserted · {errores} errores")


def _sync_pro(supabase, wb):
    """Sincroniza contratos_prorrogas desde hoja BD_CTO_PRO."""
    print("  · BD_CTO_PRO → contratos_prorrogas")
    if 'BD_CTO_PRO' not in wb.sheetnames:
        print("  ⚠ Hoja BD_CTO_PRO no encontrada")
        return

    filas = _sheet_rows(wb['BD_CTO_PRO'])
    if not filas:
        print("    · Sin datos de prórrogas")
        return

    count = errores = 0
    for row in filas:
        numero = _to_int(row.get('no.') or row.get('no'))
        if numero is None:
            continue
        data = {
            'contrato_id': CONTRATO_ID,
            'numero':      numero,
            'plazo_dias':  _to_int(row.get('plazo')),
            'fecha_fin':   _to_date(row.get('fecha_fin')),
            'fecha_firma': _to_date(row.get('fecha_firma')),
        }
        data = {k: v for k, v in data.items() if v is not None}
        try:
            supabase.table('contratos_prorrogas').upsert(
                data, on_conflict='contrato_id,numero'
            ).execute()
            count += 1
        except Exception as e:
            errores += 1
            print(f"    ✗ Prórroga #{numero}: {e}")

    print(f"    → {count} upserted · {errores} errores")


def _sync_adi(supabase, wb):
    """Sincroniza contratos_adiciones desde hoja BD_CTO_ADI."""
    print("  · BD_CTO_ADI → contratos_adiciones")
    if 'BD_CTO_ADI' not in wb.sheetnames:
        print("  ⚠ Hoja BD_CTO_ADI no encontrada")
        return

    filas = _sheet_rows(wb['BD_CTO_ADI'])
    if not filas:
        print("    · Sin datos de adiciones")
        return

    count = errores = 0
    for row in filas:
        numero = _to_int(row.get('no.') or row.get('no'))
        if numero is None:
            continue
        data = {
            'contrato_id':  CONTRATO_ID,
            'numero':       numero,
            'adicion':      _to_int(row.get('adicion')),
            'valor_actual': _to_int(row.get('valor_actual')),
            'fecha_firma':  _to_date(row.get('fecha_firma')),
        }
        data = {k: v for k, v in data.items() if v is not None}
        try:
            supabase.table('contratos_adiciones').upsert(
                data, on_conflict='contrato_id,numero'
            ).execute()
            count += 1
        except Exception as e:
            errores += 1
            print(f"    ✗ Adición #{numero}: {e}")

    print(f"    → {count} upserted · {errores} errores")
